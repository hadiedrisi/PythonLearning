#pypi.org
import openpyxl as xl

wb = xl.load_workbook("test.xlsx")
# wb['Sheet1'].title = "test"
# wb.save("./public/test.xlsx")
sheet = wb['Sheet1']
cell = sheet['A1']
cell2 = sheet.cell(row=1, column=2)
print(cell.value)
print(cell2.value)

for row in range(1, sheet.max_row+1):
    for col in range(1, sheet.max_column+1):
        cell = sheet.cell(row, col)
        print(cell.value)